"""
transcriptionCheck.py — run the transcription content checks on any delimited
or Excel file.

Standalone: no database, no email, no bucket. Point it at a file (or a folder of
files), name the column holding the transcription, and it writes an xlsx report
next to your chosen output directory.

    python transcriptionCheck.py                        # uses [INPUT] path
    python transcriptionCheck.py --input myFile.tsv
    python transcriptionCheck.py --input someFolder/ --text-column transcription
    python transcriptionCheck.py --input a.tsv --print-only

Checks run (toggle each under [CHECKS] in config.cfg):
    transcript_is_empty            blank / whitespace-only text
    number_of_words_lt_threshold   fewer than check_words_limit words
    transcript_has_numbers         digits left in the text
    brackets_dont_have_open_close  an unclosed [ ], < > or { }
    bracket_structure_invalid      malformed / crossed tag nesting
    consecutive_duplicate_tag      the same tag twice in a row
    tag_spacing_missing            a tag glued to a word with no space
    script_inconsistency           more than one Indic script in one line
    inaudible / unintelligible     a bare tag not wrapped in [ ]
    PAUSE / UNKNOWN_SEGMENT        a bare tag not wrapped in < >
    nonnative_charecters           Latin letters outside brackets
    native_script_mismatch         a character outside the language's script
    batch_minimum_acceptance       too large a share of rows failed

Exit code is 0 when nothing is flagged, 1 when any ERROR is raised.
"""

import argparse
import configparser
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from contentChecks import Finding, run_content_checks


HERE = Path(__file__).resolve().parent
DEFAULT_CONFIG = HERE / "config.cfg"
ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
READABLE_SUFFIXES = (".tsv", ".csv", ".txt", ".xlsx", ".xls")


def load_config(path: Path) -> configparser.ConfigParser:
    if not path.exists():
        raise SystemExit(f"Config not found: {path}")
    cp = configparser.ConfigParser()
    cp.read(path)
    return cp


def read_table(path: Path, sep: str) -> pd.DataFrame:
    """Read a delimited or Excel file, everything as text so nothing is
    silently coerced (a transcription of '123' must stay a string)."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(path, dtype=str).fillna("")
    return pd.read_csv(path, sep=sep, dtype=str, keep_default_na=False)


def collect_inputs(target: Path, recursive: bool = False) -> List[Path]:
    """Every readable file under `target`. A folder yields one report sheet per
    file, so a whole delivery can be checked in one run."""
    if target.is_file():
        return [target]
    if target.is_dir():
        walk = target.rglob("*") if recursive else target.iterdir()
        return sorted(p for p in walk
                      if p.is_file() and p.suffix.lower() in READABLE_SUFFIXES)
    raise SystemExit(f"Input not found: {target}")


# --------------------------------------------------------------------------- #
# Report
# --------------------------------------------------------------------------- #

def findings_to_df(findings: List[Finding]) -> pd.DataFrame:
    cols = ["severity", "check", "row", "key", "column", "message"]
    if not findings:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame([{
        "severity": f.severity, "check": f.check, "row": f.row,
        "key": f.key, "column": f.column, "message": f.message,
    } for f in findings])


def summary_row(name: str, findings: List[Finding]) -> Dict:
    n_err = sum(1 for f in findings if f.severity == "ERROR")
    n_warn = sum(1 for f in findings if f.severity == "WARNING")
    status = "ERROR" if n_err else ("WARNING" if n_warn else "OK")
    return dict(file=name, status=status, errors=n_err, warnings=n_warn,
                total=len(findings))


def clean_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Hand-typed text can carry control characters that openpyxl rejects."""
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(
                lambda v: ILLEGAL_XLSX_RE.sub(" ", v) if isinstance(v, str) else v)
    return df


def autosize(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = max((len(str(c.value)) for c in ws[letter] if c.value),
                      default=0)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 80)


def format_workbook(writer) -> None:
    fills = {
        "ERROR":   PatternFill("solid", fgColor="FCE4E4"),
        "WARNING": PatternFill("solid", fgColor="FFF2CC"),
        "OK":      PatternFill("solid", fgColor="E2F0D9"),
        "INFO":    PatternFill("solid", fgColor="DEEBF7"),
    }
    for sheet in writer.book.sheetnames:
        ws = writer.book[sheet]
        for cell in ws[1]:
            cell.font = Font(bold=True)
        status_idx = next((i for i, c in enumerate(ws[1], start=1)
                           if c.value in ("severity", "status")), None)
        if status_idx:
            for row in ws.iter_rows(min_row=2):
                fill = fills.get(row[status_idx - 1].value)
                if fill:
                    for c in row:
                        c.fill = fill
        autosize(ws)


def build_sheet_names(file_names: List[str]) -> Dict[str, Dict[str, str]]:
    """file name -> {'findings': sheet, 'rows': sheet}.

    Excel caps sheet names at 31 chars and rejects []:*?/\\. Delivery files
    routinely differ only in their last few characters (..._V1 / _V2 / _V3), so
    a plain head-truncation gives every file the same sheet name and pandas
    silently writes them all onto one sheet. Number each file and keep the TAIL
    of its stem, which is the part that actually distinguishes them.
    """
    names: Dict[str, Dict[str, str]] = {}
    for i, file_name in enumerate(file_names, start=1):
        stem = re.sub(r"[\[\]:*?/\\]", "_", Path(file_name).stem)
        prefix = f"{i}_"

        def fit(suffix: str) -> str:
            room = 31 - len(prefix) - len(suffix)
            return prefix + stem[-room:] + suffix

        names[file_name] = {"findings": fit(""), "rows": fit("_rows")}
    return names


def write_report(summary_rows: List[Dict],
                 findings_by_file: Dict[str, List[Finding]],
                 row_dfs: Dict[str, pd.DataFrame], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    sheets = build_sheet_names(list(findings_by_file))
    summary = pd.DataFrame(summary_rows)
    # Sheet names are numbered and tail-trimmed, so name the sheet each row
    # belongs to rather than leaving the reader to guess.
    if not summary.empty:
        summary.insert(1, "sheet",
                       summary["file"].map(lambda f: sheets.get(f, {}).get("findings", "")))
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        for name, findings in findings_by_file.items():
            findings_to_df(findings).to_excel(
                writer, sheet_name=sheets[name]["findings"], index=False)
        for name, df in row_dfs.items():
            if df is None or df.empty:
                continue
            clean_for_excel(df).to_excel(
                writer, sheet_name=sheets[name]["rows"], index=False)
        format_workbook(writer)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Run the transcription content checks on a file "
                    "or a folder of files.")
    ap.add_argument("--config", default=str(DEFAULT_CONFIG))
    ap.add_argument("--input", default=None,
                    help="File or folder to check (default: [INPUT] path).")
    ap.add_argument("--text-column", default=None,
                    help="Column holding the transcription.")
    ap.add_argument("--language-column", default=None,
                    help="Column holding the language name.")
    ap.add_argument("--key-column", default=None,
                    help="Column used to identify a row in the findings.")
    ap.add_argument("--sep", default=None,
                    help="Delimiter for text files (default: tab).")
    ap.add_argument("--out", default=None, help="Output xlsx path.")
    ap.add_argument("--recursive", action="store_true", default=None,
                    help="Search sub-folders too when --input is a folder.")
    ap.add_argument("--print-only", action="store_true",
                    help="Print findings, write no report.")
    args = ap.parse_args()

    cfg = load_config(Path(args.config))
    target = args.input or cfg.get("INPUT", "path", fallback="").strip()
    if not target:
        raise SystemExit("No input: pass --input or set [INPUT] path in the config.")
    text_column = args.text_column or cfg.get("INPUT", "text_column",
                                              fallback="transcription")
    language_column = args.language_column or cfg.get("INPUT", "language_column",
                                                      fallback="language")
    key_column = args.key_column or cfg.get("INPUT", "key_column",
                                            fallback="File_Name")
    sep = args.sep or cfg.get("INPUT", "sep", fallback="\t")
    sep = sep.replace("\\t", "\t")
    condition_column = cfg.get("INPUT", "condition_column", fallback="").strip()
    condition_value = cfg.get("INPUT", "condition_value", fallback="").strip()
    skip_blank = cfg.getboolean("INPUT", "skip_blank_rows", fallback=True)
    blank_values = {v.strip().lower() for v in
                    cfg.get("INPUT", "blank_values", fallback="").split(",")
                    if v.strip()}

    recursive = (args.recursive if args.recursive is not None
                 else cfg.getboolean("INPUT", "recursive", fallback=False))
    inputs = collect_inputs(Path(target).expanduser().resolve(), recursive)
    if not inputs:
        raise SystemExit(f"No readable files in {target}")

    print(f"input      : {target}")
    print(f"text column: {text_column}")
    print(f"files      : {len(inputs)}")

    summary_rows: List[Dict] = []
    findings_by_file: Dict[str, List[Finding]] = {}
    row_dfs: Dict[str, pd.DataFrame] = {}

    for path in inputs:
        print(f"\n=== {path.name} ===")
        try:
            df = read_table(path, sep)
        except Exception as e:
            findings = [Finding("ERROR", "file_read_error",
                                message=f"{type(e).__name__}: {e}")]
            findings_by_file[path.name] = findings
            summary_rows.append(summary_row(path.name, findings))
            print(f"  FAILED to read: {e}")
            continue

        keep = pd.Series(True, index=df.index)
        if condition_column:
            if condition_column in df.columns:
                # Only rows whose answer is NOT condition_value carry a
                # correction worth validating.
                keep &= (df[condition_column].astype(str).str.strip().str.lower()
                         != condition_value.lower())
                print(f"  {int((~keep).sum())} row(s) skipped: "
                      f"'{condition_column}' == '{condition_value}'")
            else:
                print(f"  WARNING: condition column '{condition_column}' not in "
                      f"this file; every row will be checked")
        if skip_blank and text_column in df.columns:
            filled = df[text_column].map(
                lambda v: str(v).strip() != ""
                and str(v).strip().lower() not in blank_values)
            n_blank = int((keep & ~filled).sum())
            keep &= filled
            if n_blank:
                print(f"  {n_blank} blank row(s) skipped")
        checked = df[keep].copy()

        findings = run_content_checks(checked, cfg, text_column,
                                      language_column, key_column)
        # Carry the flags back onto the full frame so the rows sheet lines up
        # with the source file, blank rows included.
        for col in checked.columns:
            if col.endswith("_error"):
                df[col] = False
                df.loc[checked.index, col] = checked[col]

        n_err = sum(1 for f in findings if f.severity == "ERROR")
        print(f"  {len(checked)} row(s) checked, {n_err} error(s)")
        by_check: Dict[str, int] = {}
        for f in findings:
            if f.severity == "ERROR":
                by_check[f.check] = by_check.get(f.check, 0) + 1
        for check, n in sorted(by_check.items(), key=lambda kv: -kv[1]):
            print(f"    {check:<34} {n}")

        findings_by_file[path.name] = findings
        summary_rows.append(summary_row(path.name, findings))
        row_dfs[path.name] = df

    total_err = sum(r["errors"] for r in summary_rows)
    # Width from the actual names: these files often differ only in their last
    # characters, so a fixed truncation would print identical-looking rows.
    name_w = max([len(r["file"]) for r in summary_rows] + [len("file")])
    line_w = name_w + 2 + 10 + 8 + 10
    print("\n" + "=" * line_w)
    print(f"{'file':<{name_w}}  {'status':<10}{'errors':>8}{'warnings':>10}")
    for r in summary_rows:
        print(f"{r['file']:<{name_w}}  {r['status']:<10}{r['errors']:>8}"
              f"{r['warnings']:>10}")
    print("=" * line_w)
    print(f"RESULT: {'FAIL' if total_err else 'PASS'}  ({total_err} error(s))")

    if not args.print_only:
        if args.out:
            out = Path(args.out)
        else:
            out_dir = Path(cfg.get("OUTPUT", "dir", fallback=str(HERE / "results")))
            if not out_dir.is_absolute():
                out_dir = (HERE / out_dir).resolve()
            ts = datetime.now().strftime("%d-%m-%y-%H-%M-%S")
            out = out_dir / f"TranscriptionCheckReport_{ts}.xlsx"
        write_report(summary_rows, findings_by_file, row_dfs, out)
        print(f"report -> {out}")

    return 1 if total_err else 0


if __name__ == "__main__":
    sys.exit(main())
